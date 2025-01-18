from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import TextAxis
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
import openpyxl

data_list = [0.000, 0.360, 0.753, 0.546, 0.984, 0.233, 0.531, 1.239, 1.833, 1.539, 2.043, 2.197, 0.000]
wb = Workbook()
ws = wb.worksheets[0]

# Put data in Excel
col_i= 5
for value in data_list:
    ws[f'{get_column_letter(col_i)}1'].value = value
    col_i+= 1

c = LineChart()
c.add_data(Reference(ws, min_col=5, min_row=1, max_col=5+len(data_list), max_row=1), from_rows=True)
c.graphical_properties = GraphicalProperties()

# Remove horizontal grid lines
c.y_axis.majorGridlines = None

# Remove borders
# make the border area around the plot area transparent
c.graphical_properties.noFill = True
# remove the border around the edge of the chart
c.graphical_properties.line.noFill = True

# line thickness
c.graphical_properties.line.width = pixels_to_EMU(1) # width in EMUs

# remove legend
c.legend = None
c .varyColors = False

# size
c.height = 5
c.width = 25


ws.add_chart(c, 'A2')
wb.save(r'C:\Users\agori\Python\python-tests\openpyxl_test.xlsx')
